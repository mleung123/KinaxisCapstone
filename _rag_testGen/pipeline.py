from __future__ import annotations

import json
import os
import time
import psycopg
import sys
import re

from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from openpyxl import Workbook

from embed_lmstudio import EmbedConfig, embed_texts
from ingest import IngestConfig, ingest_domain
from lmstudio_client import call_llm

from db_pgvector import (
    chunks_rowcount,
    clear_chunks,
    ensure_schema,
    get_db_snapshot_per_doc,
    get_db_snapshot_summary,
    get_random_chunks,
    similarity_search,
)

from text_utils import (
    clean_generator_text,
    enforce_hygiene_on_review,
    extract_first_json_obj,
    hard_trim_after_difficulty,
    validate_generator_schema,
)


# ---------------------------------------------------------------------
# Config dataclasses
# ---------------------------------------------------------------------


@dataclass(frozen=True)
class GenerateConfig:
    """note: Parameters for generate-only runs that read existing embedded chunks from Postgres+pgvector."""

    db_dsn: str
    lm_url: str
    embed_model: str
    generator_model: str
    context_model: str
    review_model: str
    n_items: int
    run_id: str
    prompts_dir: Path
    out_dir: Path
    top_k: int = 6
    temperature_gen: float = 0.2
    temperature_context: float = 0.0
    temperature_review: float = 0.0
    max_tokens_gen: int = 300
    max_tokens_context: int = 500
    max_tokens_review: int = 600
    request_timeout_seconds: int = 600
    sleep_seconds: float = 0.0


@dataclass(frozen=True)
class PipelineConfig:
    """note: Flat pipeline config matching cli.py construction for the 'pipeline' command."""

    db_dsn: str
    domain_dir: Path
    lm_url: str
    embed_model: str
    embedding_dim: int | None = None
    batch_size: int = 32
    chunk_chars: int = 1600
    overlap_chars: int = 200
    clear_first: bool = False
    force_ingest: bool = False
    ingest_only: bool = False
    n_items: int = 5
    generator_model: str = ""
    context_model: str = ""
    review_model: str = ""
    run_id: str = ""
    prompts_dir: Path = Path("_prompts")
    out_dir: Path = Path("runs")
    top_k: int = 6
    sleep_seconds: float = 0.0


# ---------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------


def _parse_item_fields(gen_text: str) -> dict[str, str]:
    """note: Extracts question/choices/correct_key/difficulty from the clean generator text using simple label heuristics."""
    t = gen_text or ""
    lines = [ln.rstrip() for ln in t.splitlines()]

    def _starts_with_any(s: str, prefixes: list[str]) -> bool:
        ss = (s or "").lstrip().lower()
        return any(ss.startswith(p) for p in prefixes)

    def _grab_label_value(label: str) -> str:
        lab = label.lower()
        for ln in lines:
            s = (ln or "").lstrip()
            if s.lower().startswith(lab):
                if ":" in s:
                    return s.split(":", 1)[1].strip()
                return ""
        return ""

    def _grab_multiline_after_label(label: str) -> str:
        lab = label.lower()
        i_label = None
        for i, ln in enumerate(lines):
            s = (ln or "").lstrip()
            if s.lower().startswith(lab):
                i_label = i
                if ":" in s:
                    inline = s.split(":", 1)[1].strip()
                    if inline:
                        return inline
                break

        if i_label is None:
            return ""

        parts: list[str] = []
        terminators = ["a)", "b)", "c)", "d)", "correct", "difficulty"]
        for j in range(i_label + 1, len(lines)):
            s = (lines[j] or "").strip()
            if not s:
                if parts:
                    break
                continue
            if _starts_with_any(s, terminators):
                break
            parts.append(s)

        return " ".join(parts).strip()

    out: dict[str, str] = {}
    out["question"] = _grab_multiline_after_label("question")

    for opt in ["a)", "b)", "c)", "d)"]:
        val = ""
        for ln in lines:
            s = (ln or "").lstrip()
            if s.lower().startswith(opt):
                val = s.split(")", 1)[1].strip() if ")" in s else ""
                break
        out[opt[0]] = val

    out["correct_key"] = _grab_label_value("correct_key") or _grab_label_value("correct key")
    out["difficulty"] = (_grab_label_value("difficulty") or "").lower()

    return out


_ILLEGAL_XLSX_CHARS = re.compile(r"[--]")


def _sanitize_cell(v: Any) -> Any:
    """note: Strips control characters that openpyxl rejects (e.g. ,  from python-pptx)."""
    if v is None:
        return ""
    if isinstance(v, str):
        return _ILLEGAL_XLSX_CHARS.sub(" ", v)
    return v


def _xlsx_write_sheet(ws, headers: list[str], rows: Iterable[list[Any]]) -> None:
    """note: Writes headers + rows to an openpyxl worksheet."""
    ws.append(headers)
    for r in rows:
        ws.append([_sanitize_cell(v) for v in r])


def write_run_xlsx(
    out_dir: Path,
    run_id: str,
    metadata: dict[str, Any],
    db_snapshot_summary: dict[str, Any] | None = None,
    db_snapshot_per_doc: list[dict[str, Any]] | None = None,
    chunk_preview_rows: list[dict[str, Any]] | None = None,
    items_rows: list[dict[str, Any]] | None = None,
    decisions_rows: list[dict[str, Any]] | None = None,
    trace_rows: list[dict[str, Any]] | None = None,
    include_generation: bool = True,
    include_chunk_preview: bool = True,
) -> Path:
    """note: Creates a single XLSX per run. Generation sheets are omitted in ingest-only mode."""
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / f"run_{run_id}.xlsx"

    wb = Workbook()

    # ---- Run Metadata (always first / active sheet) ----
    ws_meta = wb.active
    ws_meta.title = "Run Metadata"
    meta_headers = ["key", "value"]
    meta_rows: list[list[Any]] = []
    for k in sorted(metadata.keys()):
        v = metadata[k]
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False)
        meta_rows.append([k, v])
    _xlsx_write_sheet(ws_meta, meta_headers, meta_rows)

    # ---- DB Snapshot (always present) ----
    ws_snap = wb.create_sheet("DB Snapshot")
    ws_snap.append(["--- Summary ---", ""])
    for k, v in sorted((db_snapshot_summary or {}).items()):
        ws_snap.append([k, ("" if v is None else v)])
    ws_snap.append(["", ""])
    ws_snap.append(["--- Per-Document Inventory ---", ""])
    per_doc_headers = ["doc_path", "chunk_count", "doc_sha256", "first_created_at", "last_updated_at"]
    ws_snap.append(per_doc_headers)
    for row in (db_snapshot_per_doc or []):
        ws_snap.append([
            row.get("doc_path", ""),
            row.get("chunk_count", ""),
            row.get("doc_sha256", ""),
            row.get("first_created_at", "") or "",
            row.get("last_updated_at", "") or "",
        ])

    # ---- Chunk Preview (always present unless explicitly disabled) ----
    if include_chunk_preview:
        ws_cp = wb.create_sheet("Chunk Preview")
        cp_headers = [
            "doc_name", "chunk_index", "chunk_chars",
            "alpha_chars", "alpha_ratio", "passed_filter", "chunk_text",
        ]
        cp_data: list[list[Any]] = []
        for r in (chunk_preview_rows or []):
            cp_data.append([
                r.get("doc_name"),
                r.get("chunk_index"),
                r.get("chunk_chars"),
                r.get("alpha_chars"),
                r.get("alpha_ratio"),
                r.get("passed_filter"),
                r.get("chunk_text"),
            ])
        _xlsx_write_sheet(ws_cp, cp_headers, cp_data)

    # ---- Generation sheets (skipped in ingest-only mode) ----
    if include_generation:
        ws_items = wb.create_sheet("Items")
        items_headers = [
            "run_id", "item_id",
            "question", "a", "b", "c", "d",
            "correct_key", "difficulty",
            "decision",
            "source_alignment", "distractor_quality", "stem_clarity", "difficulty_match",
            "schema_ok", "schema_violations",
            "reviewer_schema_ok", "reviewer_schema_violations",
            "gen_text_clean",
            "seed_doc_path",
        ]
        items_data: list[list[Any]] = []
        for r in (items_rows or []):
            items_data.append([
                r.get("run_id"), r.get("item_id"),
                r.get("question"), r.get("a"), r.get("b"), r.get("c"), r.get("d"),
                r.get("correct_key"), r.get("difficulty"),
                r.get("decision"),
                r.get("source_alignment"), r.get("distractor_quality"),
                r.get("stem_clarity"), r.get("difficulty_match"),
                r.get("schema_ok"), r.get("schema_violations"),
                r.get("reviewer_schema_ok"), r.get("reviewer_schema_violations"),
                r.get("gen_text_clean"),
                r.get("seed_doc_path"),
            ])
        _xlsx_write_sheet(ws_items, items_headers, items_data)

        ws_rev = wb.create_sheet("Reviewer Decisions")
        rev_headers = [
            "run_id", "item_id", "decision",
            "source_alignment", "distractor_quality", "stem_clarity", "difficulty_match",
            "failure_layer", "reason_codes", "revision_instructions",
            "reviewer_schema_ok", "reviewer_schema_violations", "reviewer_parse_ok",
        ]
        rev_data: list[list[Any]] = []
        for r in (decisions_rows or []):
            rev_data.append([
                r.get("run_id"), r.get("item_id"), r.get("decision"),
                r.get("source_alignment"), r.get("distractor_quality"),
                r.get("stem_clarity"), r.get("difficulty_match"),
                r.get("failure_layer"),
                json.dumps(r.get("reason_codes", []), ensure_ascii=False),
                r.get("revision_instructions"),
                r.get("reviewer_schema_ok"),
                json.dumps(r.get("reviewer_schema_violations", []), ensure_ascii=False),
                r.get("reviewer_parse_ok"),
            ])
        _xlsx_write_sheet(ws_rev, rev_headers, rev_data)

        ws_trace = wb.create_sheet("Traceability")
        trace_headers = ["run_id", "item_id", "doc_path", "chunk_index", "distance", "chunk_text"]
        trace_data: list[list[Any]] = []
        for r in (trace_rows or []):
            trace_data.append([
                r.get("run_id"), r.get("item_id"), r.get("doc_path"),
                r.get("chunk_index"), r.get("distance"), r.get("chunk_text"),
            ])
        _xlsx_write_sheet(ws_trace, trace_headers, trace_data)

        ws_q = wb.create_sheet("Quality Metrics")
        ws_q.append(["metric", "value"])
        for k in sorted(metadata.keys()):
            if str(k).startswith("quality."):
                v = metadata[k]
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False)
                ws_q.append([k, ("" if v is None else v)])

    wb.save(xlsx_path)
    return xlsx_path


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------


def _redact_dsn(dsn: str) -> str:
    """note: Redacts password in a Postgres DSN for logging/metadata."""
    if not dsn:
        return ""
    try:
        if "://" in dsn and "@" in dsn:
            left, right = dsn.split("://", 1)
            creds_host = right
            if "@" in creds_host and ":" in creds_host.split("@", 1)[0]:
                creds, host = creds_host.split("@", 1)
                user = creds.split(":", 1)[0]
                return f"{left}://{user}:***@{host}"
    except Exception:
        pass
    return dsn


def _infer_embedding_dim_from_db(conn: psycopg.Connection) -> int:
    """note: Infers embedding dimension from an existing rag_chunks row."""
    with conn.cursor() as cur:
        cur.execute("SELECT embedding::text FROM rag_chunks LIMIT 1;")
        row = cur.fetchone()
    if not row or not row[0]:
        raise RuntimeError("Cannot infer embedding_dim: rag_chunks is empty.")
    s = str(row[0]).strip()
    if s.startswith("[") and s.endswith("]"):
        inner = s[1:-1].strip()
        if not inner:
            raise RuntimeError("Cannot infer embedding_dim: empty embedding literal.")
        return len([x for x in inner.split(",") if x.strip() != ""])
    raise RuntimeError(f"Cannot infer embedding_dim from embedding::text: {s[:80]}")


def _chunk_alpha_stats(text: str) -> tuple[int, int, float]:
    """note: Returns (total_chars, alpha_chars, alpha_ratio) for a chunk text."""
    t = (text or "").strip()
    total = len(t)
    alpha = sum(1 for c in t if c.isalpha())
    ratio = round(alpha / total, 3) if total > 0 else 0.0
    return total, alpha, ratio


def _build_chunk_preview_rows(conn: psycopg.Connection, min_alpha_chars: int = 120, min_alpha_ratio: float = 0.35) -> list[dict[str, Any]]:
    """note: Reads all chunks from DB and builds Chunk Preview rows including passed_filter flag."""
    with conn.cursor() as cur:
        cur.execute("SELECT doc_path, chunk_index, chunk_text FROM rag_chunks ORDER BY doc_path, chunk_index;")
        rows = cur.fetchall() or []

    preview: list[dict[str, Any]] = []
    for doc_path, chunk_index, chunk_text in rows:
        total, alpha, ratio = _chunk_alpha_stats(chunk_text or "")
        passed = (alpha >= min_alpha_chars) and (ratio >= min_alpha_ratio)
        preview.append({
            "doc_name": Path(str(doc_path)).name,
            "chunk_index": int(chunk_index),
            "chunk_chars": total,
            "alpha_chars": alpha,
            "alpha_ratio": ratio,
            "passed_filter": passed,
            "chunk_text": str(chunk_text or ""),
        })
    return preview


# ---------------------------------------------------------------------
# Generation (retrieval + agentic review)
# ---------------------------------------------------------------------


def _is_degenerate_context(text: str, max_repeat_ratio: float = 0.25, max_nonascii_ratio: float = 0.15) -> bool:
    """note: Returns True if the context agent output looks like a hallucination/degeneration loop.

    Two signals:
    1. Repeated-token ratio: split on whitespace, count how often the single most
       common token appears. If it dominates more than max_repeat_ratio of all
       tokens the model is probably looping (e.g. 'beneficiation beneficiation ...').
    2. Non-ASCII ratio: if more than max_nonascii_ratio of characters are outside
       the printable ASCII range the model has gone off-script (e.g. CJK characters,
       random unicode from a confused tokenizer).
    """
    if not text or not text.strip():
        return False

    # Signal 1: repeated token loop
    tokens = text.split()
    if tokens:
        from collections import Counter
        most_common_count = Counter(tokens).most_common(1)[0][1]
        if most_common_count / len(tokens) > max_repeat_ratio:
            return True

    # Signal 2: non-ASCII character flood
    non_ascii = sum(1 for c in text if ord(c) > 127)
    if non_ascii / len(text) > max_nonascii_ratio:
        return True

    return False

def generate_from_db(cfg: GenerateConfig) -> dict[str, Any]:
    """note: Generates items by retrieving top-k chunks, running context rewrite + generator + reviewer, writing XLSX."""
    prompts_dir = Path(cfg.prompts_dir)
    generator_system = (prompts_dir / "generator_system.txt").read_text(encoding="utf-8")
    generator_user_template = (prompts_dir / "generator_user.txt").read_text(encoding="utf-8")
    reviewer_system = (prompts_dir / "reviewer_system.txt").read_text(encoding="utf-8")
    reviewer_user_template = (prompts_dir / "reviewer_user.txt").read_text(encoding="utf-8")
    context_system = (prompts_dir / "context_system.txt").read_text(encoding="utf-8")
    context_user_template = (prompts_dir / "context_user.txt").read_text(encoding="utf-8")

    include_chunk_preview = (os.environ.get("CHUNK_PREVIEW") or "1").strip() not in {"0", "false", "FALSE", "no", "NO"}
    condition_label = (os.environ.get("CONDITION_LABEL") or "").strip() or "baseline"

    created_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    out_dir = Path(cfg.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    manifest = {
        "created_at": created_at,
        "config": {
            "db_dsn_redacted": _redact_dsn(cfg.db_dsn),
            "lm_url": cfg.lm_url,
            "embed_model": cfg.embed_model,
            "generator_model": cfg.generator_model,
            "context_model": cfg.context_model,
            "review_model": cfg.review_model,
            "n_items": int(cfg.n_items),
            "run_id": cfg.run_id,
            "prompts_dir": str(cfg.prompts_dir),
            "out_dir": str(cfg.out_dir),
            "top_k": int(cfg.top_k),
            "sleep_seconds": float(cfg.sleep_seconds),
            "condition_label": str(condition_label),
        },
        "prompt_files": {
            "generator_system": "generator_system.txt",
            "generator_user": "generator_user.txt",
            "reviewer_system": "reviewer_system.txt",
            "reviewer_user": "reviewer_user.txt",
            "context_system": "context_system.txt",
            "context_user": "context_user.txt",
        },
    }
    manifest_path = out_dir / f"run_manifest_{cfg.run_id}.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    items_rows: list[dict[str, Any]] = []
    decisions_rows: list[dict[str, Any]] = []
    trace_rows: list[dict[str, Any]] = []

    schema_ok_count = 0
    reviewer_json_ok = 0
    reviewer_schema_ok_count = 0
    decisions_count: dict[str, int] = {}

    elapsed_llm_context = 0.0
    elapsed_llm_gen = 0.0
    elapsed_llm_review = 0.0

    db_snap_summary: dict[str, Any] | None = None
    db_snap_per_doc: list[dict[str, Any]] | None = None
    chunk_preview: list[dict[str, Any]] | None = None

    def _mean(nums: list[int]) -> float | None:
        if not nums:
            return None
        return sum(nums) / float(len(nums))

    def _cap_text(s: str, max_chars: int) -> str:
        """note: Caps a string by characters to protect LM Studio context window."""
        s = s or ""
        return s if len(s) <= max_chars else (s[:max_chars] + "\n...[truncated]")

    with psycopg.connect(cfg.db_dsn) as conn:
        if chunks_rowcount(conn) <= 0:
            raise RuntimeError("DB has 0 chunks. Run 'ingest' (or 'pipeline') first.")

        embedding_dim = _infer_embedding_dim_from_db(conn)
        ensure_schema(conn, embedding_dim)

        for i in range(int(cfg.n_items)):
            item_id = f"item_{i+1}"
            print(f"Currently processing... {item_id}", file=sys.stderr, flush=True)

            seed_rows = get_random_chunks(conn, n=1)
            seed_text = seed_rows[0]["chunk_text"]
            seed_doc = seed_rows[0].get("doc_path", "")

            seed_emb = embed_texts(
                EmbedConfig(lm_url=cfg.lm_url, model=cfg.embed_model),
                [seed_text],
            )[0]

            retrieved = similarity_search(conn, seed_emb, int(cfg.top_k))

            for r in retrieved:
                trace_rows.append({
                    "run_id": cfg.run_id,
                    "item_id": item_id,
                    "doc_path": r.get("doc_path"),
                    "chunk_index": r.get("chunk_index"),
                    "distance": r.get("distance"),
                    "chunk_text": r.get("chunk_text"),
                })

            max_chunk_chars = int(os.environ.get("MAX_CHUNK_CHARS_IN_CONTEXT", "900"))
            max_context_chars_gen = int(os.environ.get("MAX_CONTEXT_CHARS_GEN", "2800"))
            max_context_chars_rev = int(os.environ.get("MAX_CONTEXT_CHARS_REV", "1800"))

            parts: list[str] = []
            for j, r in enumerate(retrieved):
                ch = _cap_text(str(r.get("chunk_text", "")), max_chunk_chars)
                parts.append(f"[{j+1}] {ch}")

            context_block_raw = "\n\n".join(parts)

            max_context_chars_rewrite = 1600
            rewrite_input = _cap_text(context_block_raw, max_context_chars_rewrite)
            context_user = context_user_template.replace("{{CONTEXT}}", rewrite_input)

            _t0 = time.perf_counter()
            rewritten_context = call_llm(
                lm_url=cfg.lm_url,
                model=cfg.context_model,
                system_prompt=context_system,
                user_prompt=context_user,
                temperature=cfg.temperature_context,
                max_tokens=cfg.max_tokens_context,
                request_timeout_seconds=cfg.request_timeout_seconds,
            )
            elapsed_llm_context += time.perf_counter() - _t0

            context_block_clean = rewritten_context.strip() or context_block_raw
            context_block_gen = _cap_text(context_block_clean, max_context_chars_gen)
            context_block_rev = _cap_text(context_block_clean, max_context_chars_rev)

            generator_user = generator_user_template.replace("{{CONTEXT}}", context_block_gen)

            _t0 = time.perf_counter()
            gen_raw = call_llm(
                lm_url=cfg.lm_url,
                model=cfg.generator_model,
                system_prompt=generator_system,
                user_prompt=generator_user,
                temperature=cfg.temperature_gen,
                max_tokens=cfg.max_tokens_gen,
                request_timeout_seconds=cfg.request_timeout_seconds,
            )
            elapsed_llm_gen += time.perf_counter() - _t0

            gen_text = hard_trim_after_difficulty(clean_generator_text(gen_raw))
            schema_ok, violations = validate_generator_schema(gen_text)
            if schema_ok:
                schema_ok_count += 1

            parsed_fields = _parse_item_fields(gen_text)

            reviewer_user = (
                reviewer_user_template
                .replace("{{GEN_ITEM}}", gen_text)
                .replace("{{CONTEXT}}", context_block_rev)
            )

            _t0 = time.perf_counter()
            rev_raw = call_llm(
                lm_url=cfg.lm_url,
                model=cfg.review_model,
                system_prompt=reviewer_system,
                user_prompt=reviewer_user,
                temperature=cfg.temperature_review,
                max_tokens=cfg.max_tokens_review,
                request_timeout_seconds=cfg.request_timeout_seconds,
            )
            elapsed_llm_review += time.perf_counter() - _t0

            rev_json = extract_first_json_obj(rev_raw) or {}
            rev_clean = enforce_hygiene_on_review(rev_json)

            if rev_clean.get("reviewer_parse_ok"):
                reviewer_json_ok += 1
            if rev_clean.get("reviewer_schema_ok"):
                reviewer_schema_ok_count += 1

            decision = rev_clean.get("decision", "")
            decisions_count[decision] = decisions_count.get(decision, 0) + 1

            items_rows.append({
                "run_id": cfg.run_id,
                "item_id": item_id,
                "question": parsed_fields.get("question", ""),
                "a": parsed_fields.get("a", ""),
                "b": parsed_fields.get("b", ""),
                "c": parsed_fields.get("c", ""),
                "d": parsed_fields.get("d", ""),
                "correct_key": parsed_fields.get("correct_key", ""),
                "difficulty": parsed_fields.get("difficulty", ""),
                "decision": decision,
                "source_alignment": rev_clean.get("source_alignment"),
                "distractor_quality": rev_clean.get("distractor_quality"),
                "stem_clarity": rev_clean.get("stem_clarity"),
                "difficulty_match": rev_clean.get("difficulty_match"),
                "schema_ok": bool(schema_ok),
                "schema_violations": "|".join(violations),
                "reviewer_schema_ok": bool(rev_clean.get("reviewer_schema_ok", False)),
                "reviewer_schema_violations": "|".join(rev_clean.get("reviewer_schema_violations", []) or []),
                "gen_text_clean": gen_text,
                "seed_doc_path": seed_doc,
            })

            decisions_rows.append({
                "run_id": cfg.run_id,
                "item_id": item_id,
                "decision": decision,
                "source_alignment": rev_clean.get("source_alignment"),
                "distractor_quality": rev_clean.get("distractor_quality"),
                "stem_clarity": rev_clean.get("stem_clarity"),
                "difficulty_match": rev_clean.get("difficulty_match"),
                "failure_layer": rev_clean.get("failure_layer", ""),
                "reason_codes": rev_clean.get("reason_codes", []),
                "revision_instructions": rev_clean.get("revision_instructions", ""),
                "reviewer_schema_ok": bool(rev_clean.get("reviewer_schema_ok", False)),
                "reviewer_schema_violations": rev_clean.get("reviewer_schema_violations", []),
                "reviewer_parse_ok": bool(rev_clean.get("reviewer_parse_ok", False)),
            })

            if cfg.sleep_seconds:
                time.sleep(float(cfg.sleep_seconds))

        db_snap_summary = get_db_snapshot_summary(conn)
        db_snap_per_doc = get_db_snapshot_per_doc(conn)
        if include_chunk_preview:
            chunk_preview = _build_chunk_preview_rows(conn)

    elapsed_llm_total = elapsed_llm_context + elapsed_llm_gen + elapsed_llm_review
    print(
        f"Timing: context={elapsed_llm_context:.1f}s  gen={elapsed_llm_gen:.1f}s  "
        f"review={elapsed_llm_review:.1f}s  llm_total={elapsed_llm_total:.1f}s",
        file=sys.stderr, flush=True,
    )

    valid_reviews = [r for r in decisions_rows if bool(r.get("reviewer_schema_ok"))]
    sa_vals = [int(r["source_alignment"]) for r in valid_reviews if r.get("source_alignment") is not None]
    dq_vals = [int(r["distractor_quality"]) for r in valid_reviews if r.get("distractor_quality") is not None]
    sc_vals = [int(r["stem_clarity"]) for r in valid_reviews if r.get("stem_clarity") is not None]
    dm_vals = [bool(r["difficulty_match"]) for r in valid_reviews if r.get("difficulty_match") is not None]

    def _pct(num: int, den: int) -> float | None:
        if den <= 0:
            return None
        return (float(num) / float(den)) * 100.0

    sa_gte_4 = sum(1 for v in sa_vals if v >= 4)
    dq_gte_3 = sum(1 for v in dq_vals if v >= 3)
    sc_gte_4 = sum(1 for v in sc_vals if v >= 4)
    dm_true = sum(1 for v in dm_vals if v is True)

    meta = {
        "created_at": created_at,
        "run_id": cfg.run_id,
        "condition_label": str(condition_label),
        "lm_url": cfg.lm_url,
        "embed_model": cfg.embed_model,
        "generator_model": cfg.generator_model,
        "context_model": cfg.context_model,
        "review_model": cfg.review_model,
        "n_items": int(cfg.n_items),
        "top_k": int(cfg.top_k),
        "sleep_seconds": float(cfg.sleep_seconds),
        "db_dsn_redacted": _redact_dsn(cfg.db_dsn),
        "prompts_dir": str(cfg.prompts_dir),
        "out_dir": str(cfg.out_dir),
        "quality.valid_review_rows": int(len(valid_reviews)),
        "quality.mean_source_alignment": _mean(sa_vals),
        "quality.mean_distractor_quality": _mean(dq_vals),
        "quality.mean_stem_clarity": _mean(sc_vals),
        "quality.pct_source_alignment_gte_4": _pct(sa_gte_4, len(sa_vals)),
        "quality.pct_distractor_quality_gte_3": _pct(dq_gte_3, len(dq_vals)),
        "quality.pct_stem_clarity_gte_4": _pct(sc_gte_4, len(sc_vals)),
        "quality.pct_difficulty_match_true": _pct(dm_true, len(dm_vals)),
        "timing.llm_context_seconds": round(elapsed_llm_context, 1),
        "timing.llm_gen_seconds": round(elapsed_llm_gen, 1),
        "timing.llm_review_seconds": round(elapsed_llm_review, 1),
        "timing.llm_total_seconds": round(elapsed_llm_total, 1),
    }

    xlsx_path = write_run_xlsx(
        out_dir=Path(cfg.out_dir),
        run_id=cfg.run_id,
        metadata=meta,
        db_snapshot_summary=db_snap_summary,
        db_snapshot_per_doc=db_snap_per_doc,
        chunk_preview_rows=chunk_preview,
        items_rows=items_rows,
        decisions_rows=decisions_rows,
        trace_rows=trace_rows,
        include_generation=True,
        include_chunk_preview=include_chunk_preview,
    )

    return {
        "run_id": cfg.run_id,
        "condition_label": str(condition_label),
        "out_dir": str(cfg.out_dir),
        "items_total": int(cfg.n_items),
        "items_schema_ok": int(schema_ok_count),
        "reviewer_json_ok": int(reviewer_json_ok),
        "reviewer_schema_ok": int(reviewer_schema_ok_count),
        "decisions": decisions_count,
        "files": {
            "xlsx": str(xlsx_path),
            "manifest_json": str(manifest_path),
        },
    }


# ---------------------------------------------------------------------
# Pipeline orchestrator
# ---------------------------------------------------------------------


def run_pipeline(cfg: PipelineConfig) -> dict[str, Any]:
    """note: Orchestrates ingest-if-needed, optional ingest-only XLSX, and generate."""
    ingest_ran = False
    ingest_summary: dict[str, Any] | None = None

    with psycopg.connect(cfg.db_dsn) as conn:
        has_chunks = chunks_rowcount(conn) > 0

    if cfg.force_ingest or not has_chunks:
        ingest_ran = True
        ingest_cfg = IngestConfig(
            domain_dir=Path(cfg.domain_dir),
            db_dsn=cfg.db_dsn,
            embed_lm_url=cfg.lm_url,
            embed_model=cfg.embed_model,
            embedding_dim=int(cfg.embedding_dim) if cfg.embedding_dim is not None else None,
            batch_size=int(cfg.batch_size),
            chunk_chars=int(cfg.chunk_chars),
            overlap_chars=int(cfg.overlap_chars),
            clear_first=bool(cfg.clear_first),
        )
        ingest_summary = ingest_domain(ingest_cfg)

    # ---- Ingest-only path: write XLSX with DB Snapshot + Chunk Preview, skip generation ----
    if cfg.ingest_only:
        include_chunk_preview = (os.environ.get("CHUNK_PREVIEW") or "1").strip() not in {"0", "false", "FALSE", "no", "NO"}
        created_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        out_dir = Path(cfg.out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        run_id = cfg.run_id or datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%SZ")

        meta = {
            "created_at": created_at,
            "run_id": run_id,
            "mode": "ingest_only",
            "lm_url": cfg.lm_url,
            "embed_model": cfg.embed_model,
            "db_dsn_redacted": _redact_dsn(cfg.db_dsn),
            "domain_dir": str(cfg.domain_dir),
            "chunk_chars": int(cfg.chunk_chars),
            "overlap_chars": int(cfg.overlap_chars),
            "ingest_summary": json.dumps(ingest_summary or {}, ensure_ascii=False),
        }

        with psycopg.connect(cfg.db_dsn) as conn:
            db_snap_summary = get_db_snapshot_summary(conn)
            db_snap_per_doc = get_db_snapshot_per_doc(conn)
            chunk_preview = _build_chunk_preview_rows(conn) if include_chunk_preview else None

        manifest = {"created_at": created_at, "config": meta}
        manifest_path = out_dir / f"run_manifest_{run_id}.json"
        manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

        xlsx_path = write_run_xlsx(
            out_dir=out_dir,
            run_id=run_id,
            metadata=meta,
            db_snapshot_summary=db_snap_summary,
            db_snapshot_per_doc=db_snap_per_doc,
            chunk_preview_rows=chunk_preview,
            include_generation=False,
            include_chunk_preview=include_chunk_preview,
        )

        print(f"Ingest-only XLSX written: {xlsx_path}", file=sys.stderr, flush=True)

        return {
            "ingest_ran": ingest_ran,
            "ingest_summary": ingest_summary,
            "ingest_only": True,
            "files": {
                "xlsx": str(xlsx_path),
                "manifest_json": str(manifest_path),
            },
        }

    # ---- Full pipeline: generate ----
    gen_cfg = GenerateConfig(
        db_dsn=cfg.db_dsn,
        lm_url=cfg.lm_url,
        embed_model=cfg.embed_model,
        generator_model=cfg.generator_model,
        context_model=cfg.context_model,
        review_model=cfg.review_model,
        n_items=int(cfg.n_items),
        run_id=cfg.run_id,
        prompts_dir=Path(cfg.prompts_dir),
        out_dir=Path(cfg.out_dir),
        top_k=int(cfg.top_k),
        sleep_seconds=float(cfg.sleep_seconds),
    )

    generate_summary = generate_from_db(gen_cfg)

    return {
        "ingest_ran": ingest_ran,
        "ingest_summary": ingest_summary,
        "generate_summary": generate_summary,
    }